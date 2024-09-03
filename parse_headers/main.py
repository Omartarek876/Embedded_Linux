import openpyxl
import re
import os

def parse_hf(header_path):
    with open(header_path, "r") as file:
        content = file.read()

    # Regular expressions (excluding preprocessor directives)
    patterns = {
        'function_pattern': re.compile(r'\b[A-Za-z_][A-Za-z0-9_]*\s+\**\b[A-Za-z_][A-Za-z0-9_]*\s*\([^)]*\)\s*;'),
        'global_variable_pattern': re.compile(r'\b(?:extern\s+)?[A-Za-z_][A-Za-z0-9_]*\s+\**\b[A-Za-z_][A-Za-z0-9_]*\s*(?:=\s*[^;]+)?\s*;'),
        'macro_pattern': re.compile(r'#define\s+\w+\s+.+'),
        'include_pattern': re.compile(r'#include\s*[<"]\S*[>"]'),
        'typedef_pattern': re.compile(r'typedef\s+.*\s+\w+\s*;'),
        'struct_pattern': re.compile(r'struct\s+\w+\s*\{[^}]*\}\s*;'),
        'enum_pattern': re.compile(r'enum\s+\w*\s*\{([^}]*)\}\s*;'),
        'function_like_macro_pattern': re.compile(r'#define\s+\w+\(.*?\)\s+.+')
    }

    # Extract data for each pattern
    data = {name: pattern.findall(content) for name, pattern in patterns.items()}

    # Extract enum names and their values
    enums = patterns['enum_pattern'].findall(content)
    enum_names = []
    for enum in enums:
        # Split by commas and strip spaces to get individual enum values
        values = [val.strip() for val in enum.split(',') if val.strip()]
        enum_names.extend(values)
    data['enum_pattern'] = enum_names

    return data

def append_data(sheet, data, type_label):
    if data:
        # Add data to the sheet
        for i, item in enumerate(data, start=1):
            print(f"Appending {type_label}: ID={i}, {item}")
            sheet.append([i, item])
    else:
        print(f"No data found for {type_label}. Appending 'not mentioned in the header file'.")
        sheet.append([1, 'not mentioned in the header file'])

def create_excel(out_file):
    res = 'n'
    while res == 'n':
        if os.path.exists(out_file):
            res = input(f"{out_file} already exists. [y to remove / n to change the file name]: ")
            if res == 'y':
                os.remove(out_file)
                print(f"{out_file} has been deleted and a new one will be created.")
            elif res == 'n':
                out_file = input("Enter the new file name: ") + ".xlsx"
                continue
            else:
                print("Invalid input")
                continue
        else:
            print(f"{out_file} does not exist and your file will be created.")
            break

    # Create a new workbook
    workbook = openpyxl.Workbook()

    # Create the main sheet for functions
    functions_sheet = workbook.active
    functions_sheet.title = "Functions"
    functions_sheet.append(["ID", "Function Prototype"])

    # Create sheets for other patterns
    sheets = {
        'Global Variables': workbook.create_sheet(title="Global Variables"),
        'Macros': workbook.create_sheet(title="Macros"),
        'Includes': workbook.create_sheet(title="Includes"),
        'Typedefs': workbook.create_sheet(title="Typedefs"),
        'Structs': workbook.create_sheet(title="Structs"),
        'Enums': workbook.create_sheet(title="Enums"),
        'Function-Like Macros': workbook.create_sheet(title="Function-Like Macros")
    }

    # Adjust the column width
    for sheet in [functions_sheet] + list(sheets.values()):
        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 50

    return functions_sheet, sheets, workbook

def header_path():
    while True:
        header = input("Enter the header path: ").strip()
        if os.path.exists(header):
            if os.path.isfile(header) and header.endswith('.h'):
                print("File exists and is a valid header file.")
                break
            else:
                print("The file does not have a .h extension or is not a file.")
        else:
            print("File does not exist or path is invalid.")
    return header

def main():
    header = header_path()
    out_file = input("Enter the output file name (without extension): ").strip() + ".xlsx"
    
    functions_sheet, sheets, workbook = create_excel(out_file)

    # Parse the header file
    data = parse_hf(header)
    
    # Append data to the respective sheets
    append_data(functions_sheet, data['function_pattern'], "Function Prototype")
    append_data(sheets['Global Variables'], data['global_variable_pattern'], "Global Variable Declaration")
    append_data(sheets['Macros'], data['macro_pattern'], "Macro Definition")
    append_data(sheets['Includes'], data['include_pattern'], "Include Statement")
    append_data(sheets['Typedefs'], data['typedef_pattern'], "Typedef Declaration")
    append_data(sheets['Structs'], data['struct_pattern'], "Struct Declaration")
    append_data(sheets['Enums'], data['enum_pattern'], "Enum Declaration")
    append_data(sheets['Function-Like Macros'], data['function_like_macro_pattern'], "Function-Like Macro")

    workbook.save(out_file)  # Save the workbook with the user-provided file name
    print("Done")

if __name__ == "__main__":
    main()
